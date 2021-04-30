import os
import openpyxl as xl
from openpyxl.styles import Border,Side,Alignment
from openpyxl import utils


os.chdir("python_project")#python 当前的工作目录
Path = os.getcwd()
#查看当前目录是否有'test.xlsx',没有则创建excel
ex_file = os.path.join(Path,'test.xlsx')
if not os.path.exists(ex_file):
    wb = xl.Workbook()
    worksheet = wb.active
    worksheet.title = 'Mysheet'
    wb.save(ex_file)
else:
    wb = xl.load_workbook(ex_file)
    worksheet = wb.active
    if worksheet.title == 'Mysheet':
        wb.remove(worksheet)#旧表单有数据的单元格是'ready-only',若更新表单要删除重建新表单。
        wb.create_sheet('Mysheet')
        wb.save(ex_file)

wb=xl.load_workbook("test.xlsx")
ws=wb.active
ws.column_dimensions[utils.cell.get_column_letter(1)].width=30 #设置列宽
#设置单元格格式
thin = Side(border_style="thin", color="000000")
border = Border(left = thin, right = thin, top = thin, bottom = thin)
align = Alignment(horizontal = "center",vertical = "center")
#写入数据
log=['hello']
for i in range(1,3):#openpyxl 创建的excel 第一行第一列的索引是‘A1’,cell 的行列参数 >=1。
    for j in range(1,3):
        d=ws.cell(i,j,value=log[0])
        d.border = border
        d.alignment = align

#合并单元格要从有数据单元格的起始行和起始列合并，否则会覆盖掉原始数据。合并的单元格会保留合并前的格式
d = ws.cell(3,2,"world")
d.border = border
d.alignment = align
ws.merge_cells(start_row=3,start_column=2,end_row=3,end_column=5)
#ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=5)#会覆盖掉第三行第二列的数据
wb.save("test.xlsx")
